from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.relatorio_movimento_diario_service import RelatorioMovimentoDiarioService


bp = Blueprint(
    "relatorio_movimento_diario",
    __name__,
    url_prefix="/relatorios/movimento-diario"
)


def moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def _obter_filtros():
    return {
        "data_inicio": request.args.get("data_inicio", "").strip(),
        "data_fim": request.args.get("data_fim", "").strip(),
        "unidade_id": request.args.get("unidade_id", "").strip(),
        "situacao_pagamento": request.args.get("situacao_pagamento", "").strip(),
        "origem": request.args.get("origem", "").strip(),
    }


@bp.app_template_filter("moeda")
def filtro_moeda(valor):
    return moeda(valor)


@bp.route("/", methods=["GET"])
def index():
    filtros = _obter_filtros()

    unidades = RelatorioMovimentoDiarioService.listar_unidades()

    resultado = RelatorioMovimentoDiarioService.gerar_relatorio(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        situacao_pagamento=filtros["situacao_pagamento"],
        origem=filtros["origem"],
    )

    return render_template(
        "relatorios/movimento_diario.html",
        filtros=filtros,
        unidades=unidades,
        linhas=resultado["linhas"],
        linhas_detalhe=resultado["linhas_detalhe"],
        totais=resultado["totais"],
    )


@bp.route("/exportar-excel", methods=["GET"])
def exportar_excel():
    filtros = _obter_filtros()

    resultado = RelatorioMovimentoDiarioService.gerar_relatorio(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        situacao_pagamento=filtros["situacao_pagamento"],
        origem=filtros["origem"],
    )

    linhas = resultado["linhas"]
    linhas_detalhe = resultado["linhas_detalhe"]
    totais = resultado["totais"]

    wb = Workbook()

    ws = wb.active
    ws.title = "Movimento Diário"

    cor_titulo = "1F4E78"
    cor_header = "D9EAF7"
    cor_total = "E2F0D9"
    branco = "FFFFFF"

    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:L1")
    ws["A1"] = "Relatório de Movimento Diário de Vendas"
    ws["A1"].font = Font(size=16, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Gerado em:"
    ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    ws["A4"] = "Período:"
    ws["B4"] = f"{filtros.get('data_inicio') or 'Início'} até {filtros.get('data_fim') or 'Hoje'}"

    ws["A5"] = "Situação Pagamento:"
    ws["B5"] = filtros.get("situacao_pagamento") or "Todas"

    ws["D5"] = "Origem:"
    ws["E5"] = filtros.get("origem") or "Todas"

    linha_resumo = 7

    resumo = [
        ("Dias com Movimento", totais["total_dias"]),
        ("Registros Agrupados", totais["total_linhas"]),
        ("Pedidos", totais["total_pedidos"]),
        ("Itens", totais["total_itens"]),
        ("Valor Produtos", totais["total_produtos"]),
        ("Descontos", totais["total_descontos"]),
        ("Entrega / Taxas", totais["total_entregas"]),
        ("Total Geral", totais["total_geral"]),
        ("Total Pago", totais["total_pago"]),
        ("Total Pendente", totais["total_pendente"]),
        ("Total Cancelado", totais["total_cancelado"]),
        ("Ticket Médio", totais["ticket_medio"]),
    ]

    ws[f"A{linha_resumo}"] = "Resumo Consolidado"
    ws[f"A{linha_resumo}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_resumo}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_resumo, start_column=1, end_row=linha_resumo, end_column=4)

    linha_resumo += 1

    for nome, valor in resumo:
        celula_nome = ws.cell(row=linha_resumo, column=1, value=nome)
        celula_valor = ws.cell(row=linha_resumo, column=2, value=valor)

        celula_nome.font = Font(bold=True)
        celula_nome.fill = PatternFill("solid", fgColor=cor_total)
        celula_valor.fill = PatternFill("solid", fgColor=cor_total)

        celula_nome.border = borda
        celula_valor.border = borda

        if nome not in ["Dias com Movimento", "Registros Agrupados", "Pedidos", "Itens"]:
            celula_valor.number_format = 'R$ #,##0.00'

        linha_resumo += 1

    linha_tabela = linha_resumo + 2

    ws[f"A{linha_tabela}"] = "Movimento por Dia, Origem e Filial"
    ws[f"A{linha_tabela}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_tabela}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_tabela, start_column=1, end_row=linha_tabela, end_column=12)

    linha_tabela += 1

    cabecalhos = [
        "Data",
        "Origem",
        "Filial",
        "Pedidos",
        "Itens",
        "Produtos",
        "Descontos",
        "Entrega",
        "Total",
        "Pago",
        "Pendente",
        "Ticket Médio",
    ]

    for col, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=linha_tabela, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    linha_atual = linha_tabela + 1

    for item in linhas:
        valores = [
            item.get("data_formatada"),
            item.get("origem_formatada"),
            item.get("unidade_nome"),
            item.get("quantidade_pedidos"),
            item.get("quantidade_itens"),
            item.get("valor_produtos"),
            item.get("valor_descontos"),
            item.get("valor_entregas"),
            item.get("valor_total"),
            item.get("valor_pago"),
            item.get("valor_pendente"),
            item.get("ticket_medio"),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha_atual, column=col, value=valor)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")

            if col in [6, 7, 8, 9, 10, 11, 12]:
                celula.number_format = 'R$ #,##0.00'

        linha_atual += 1

    ws_detalhe = wb.create_sheet("Detalhe dos Pedidos")

    ws_detalhe.merge_cells("A1:I1")
    ws_detalhe["A1"] = "Detalhamento dos Pedidos"
    ws_detalhe["A1"].font = Font(size=16, bold=True, color=branco)
    ws_detalhe["A1"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws_detalhe["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_detalhe.row_dimensions[1].height = 28

    linha_detalhe = 3

    cabecalhos_detalhe = [
        "Pedido",
        "Data",
        "Origem",
        "Filial",
        "Cliente",
        "Status",
        "Pagamento",
        "Itens",
        "Total",
    ]

    for col, titulo in enumerate(cabecalhos_detalhe, start=1):
        celula = ws_detalhe.cell(row=linha_detalhe, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    linha_detalhe += 1

    for item in linhas_detalhe:
        valores = [
            item.get("codigo_pedido"),
            item.get("data_formatada"),
            item.get("origem_formatada"),
            item.get("unidade_nome"),
            item.get("cliente_nome"),
            item.get("status"),
            item.get("situacao_pagamento"),
            item.get("qtd_itens"),
            item.get("valor_total"),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws_detalhe.cell(row=linha_detalhe, column=col, value=valor)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")

            if col == 9:
                celula.number_format = 'R$ #,##0.00'

        linha_detalhe += 1

    for aba in [ws, ws_detalhe]:
        for col in range(1, 13):
            letra = get_column_letter(col)
            aba.column_dimensions[letra].width = 18

        aba.column_dimensions["B"].width = 20
        aba.column_dimensions["C"].width = 28
        aba.column_dimensions["E"].width = 30

    ws.freeze_panes = "A2"
    ws_detalhe.freeze_panes = "A4"

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    nome_arquivo = f"relatorio_movimento_diario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )