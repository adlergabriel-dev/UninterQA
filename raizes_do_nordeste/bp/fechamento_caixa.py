from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, send_file, redirect, url_for, flash, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.fechamento_caixa_service import FechamentoCaixaService


bp = Blueprint(
    "fechamento_caixa",
    __name__,
    url_prefix="/relatorios/fechamento-caixa"
)


def moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def _obter_filtros():
    return {
        "data_inicio": request.args.get("data_inicio", "").strip(),
        "data_fim": request.args.get("data_fim", "").strip(),
        "unidade_id": request.args.get("unidade_id", "").strip(),
        "forma_pagamento": request.args.get("forma_pagamento", "").strip(),
        "situacao_pagamento": request.args.get("situacao_pagamento", "").strip(),
        "origem": request.args.get("origem", "").strip(),
    }


@bp.app_template_filter("moeda")
def filtro_moeda(valor):
    return moeda(valor)


@bp.route("/", methods=["GET"])
def index():
    filtros = _obter_filtros()

    unidades = FechamentoCaixaService.listar_unidades()

    resultado = FechamentoCaixaService.gerar_fechamento(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        forma_pagamento=filtros["forma_pagamento"],
        situacao_pagamento=filtros["situacao_pagamento"],
        origem=filtros["origem"],
    )

    fechamentos = FechamentoCaixaService.listar_fechamentos()

    return render_template(
        "relatorios/fechamento_caixa.html",
        filtros=filtros,
        unidades=unidades,
        linhas=resultado["linhas"],
        totais=resultado["totais"],
        fechamentos=fechamentos,
    )


@bp.route("/fechar", methods=["POST"])
def fechar():
    usuario = session.get("usuario") or {}

    data_inicio = (request.form.get("data_inicio") or "").strip()
    data_fim = (request.form.get("data_fim") or "").strip()
    unidade_id = (request.form.get("unidade_id") or "").strip()
    observacao = (request.form.get("observacao") or "").strip()

    sucesso, mensagem, fechamento_id = FechamentoCaixaService.fechar_caixa(
        data_inicio=data_inicio,
        data_fim=data_fim,
        unidade_id=unidade_id,
        usuario_id=usuario.get("id", ""),
        usuario_nome=usuario.get("nome") or usuario.get("email") or "USUARIO_INTERNO",
        observacao=observacao,
    )

    flash(mensagem, "success" if sucesso else "danger")

    return redirect(url_for(
        "fechamento_caixa.index",
        data_inicio=data_inicio,
        data_fim=data_fim,
        unidade_id=unidade_id,
    ))


@bp.route("/exportar-excel", methods=["GET"])
def exportar_excel():
    filtros = _obter_filtros()

    resultado = FechamentoCaixaService.gerar_fechamento(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        forma_pagamento=filtros["forma_pagamento"],
        situacao_pagamento=filtros["situacao_pagamento"],
        origem=filtros["origem"],
    )

    linhas = resultado["linhas"]
    totais = resultado["totais"]

    wb = Workbook()

    ws = wb.active
    ws.title = "Fechamento"

    cor_titulo = "1F4E78"
    cor_header = "D9EAF7"
    cor_total = "E2F0D9"
    branco = "FFFFFF"

    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "Fechamento de Caixa por Filial"
    ws["A1"].font = Font(size=16, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Gerado em:"
    ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    ws["A4"] = "Período:"
    ws["B4"] = f"{filtros.get('data_inicio') or 'Início'} até {filtros.get('data_fim') or 'Hoje'}"

    ws["A5"] = "Forma Pagamento:"
    ws["B5"] = filtros.get("forma_pagamento") or "Todas"

    ws["D5"] = "Situação Pagamento:"
    ws["E5"] = filtros.get("situacao_pagamento") or "Todas"

    ws["G5"] = "Origem:"
    ws["H5"] = filtros.get("origem") or "Todas"

    linha_resumo = 7

    resumo = [
        ("Total de Pedidos", totais["total_pedidos"]),
        ("Total de Itens", totais["total_itens"]),
        ("Valor Produtos", totais["total_produtos"]),
        ("Descontos", totais["total_descontos"]),
        ("Entrega / Taxas", totais["total_entregas"]),
        ("Total Geral", totais["total_geral"]),
        ("Total Pago", totais["total_pago"]),
        ("Total Pendente", totais["total_pendente"]),
        ("Total Cancelado", totais["total_cancelado"]),
    ]

    ws[f"A{linha_resumo}"] = "Resumo Consolidado"
    ws[f"A{linha_resumo}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_resumo}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_resumo, start_column=1, end_row=linha_resumo, end_column=4)

    linha_resumo += 1

    for nome, valor in resumo:
        celula_nome = ws.cell(row=linha_resumo, column=1, value=nome)
        celula_valor = ws.cell(row=linha_resumo, column=2, value=valor)

        celula_nome.font = Font(bold=True)
        celula_nome.fill = PatternFill("solid", fgColor=cor_total)
        celula_valor.fill = PatternFill("solid", fgColor=cor_total)

        celula_nome.border = borda
        celula_valor.border = borda

        if nome not in ["Total de Pedidos", "Total de Itens"]:
            celula_valor.number_format = 'R$ #,##0.00'

        linha_resumo += 1

    linha_origem = linha_resumo + 2

    ws[f"A{linha_origem}"] = "Resumo por Origem"
    ws[f"A{linha_origem}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_origem}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_origem, start_column=1, end_row=linha_origem, end_column=6)

    linha_origem += 1

    cabecalhos_origem = [
        "Origem",
        "Qtd",
        "Total",
        "Pago",
        "Pendente",
        "Cancelado",
    ]

    for col, titulo in enumerate(cabecalhos_origem, start=1):
        celula = ws.cell(row=linha_origem, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda

    linha_origem += 1

    for nome, item in totais["resumo_por_origem"].items():
        valores = [
            nome,
            item.get("quantidade", 0),
            item.get("valor_total", 0),
            item.get("valor_pago", 0),
            item.get("valor_pendente", 0),
            item.get("valor_cancelado", 0),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha_origem, column=col, value=valor)
            celula.border = borda

            if col in [3, 4, 5, 6]:
                celula.number_format = 'R$ #,##0.00'

        linha_origem += 1

    linha_forma = linha_origem + 2

    ws[f"A{linha_forma}"] = "Resumo por Forma de Pagamento"
    ws[f"A{linha_forma}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_forma}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_forma, start_column=1, end_row=linha_forma, end_column=6)

    linha_forma += 1

    cabecalhos_forma = [
        "Forma",
        "Qtd",
        "Total",
        "Pago",
        "Pendente",
        "Cancelado",
    ]

    for col, titulo in enumerate(cabecalhos_forma, start=1):
        celula = ws.cell(row=linha_forma, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda

    linha_forma += 1

    for nome, item in totais["resumo_por_forma"].items():
        valores = [
            nome,
            item.get("quantidade", 0),
            item.get("valor_total", 0),
            item.get("valor_pago", 0),
            item.get("valor_pendente", 0),
            item.get("valor_cancelado", 0),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha_forma, column=col, value=valor)
            celula.border = borda

            if col in [3, 4, 5, 6]:
                celula.number_format = 'R$ #,##0.00'

        linha_forma += 1

    linha_filial = linha_forma + 2

    ws[f"A{linha_filial}"] = "Resumo por Filial"
    ws[f"A{linha_filial}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_filial}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_filial, start_column=1, end_row=linha_filial, end_column=6)

    linha_filial += 1

    cabecalhos_filial = [
        "Filial",
        "Qtd",
        "Total",
        "Pago",
        "Pendente",
        "Cancelado",
    ]

    for col, titulo in enumerate(cabecalhos_filial, start=1):
        celula = ws.cell(row=linha_filial, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda

    linha_filial += 1

    for nome, item in totais["resumo_por_filial"].items():
        valores = [
            nome,
            item.get("quantidade", 0),
            item.get("valor_total", 0),
            item.get("valor_pago", 0),
            item.get("valor_pendente", 0),
            item.get("valor_cancelado", 0),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha_filial, column=col, value=valor)
            celula.border = borda

            if col in [3, 4, 5, 6]:
                celula.number_format = 'R$ #,##0.00'

        linha_filial += 1

    linha_detalhe = linha_filial + 2

    ws[f"A{linha_detalhe}"] = "Detalhamento dos Pedidos"
    ws[f"A{linha_detalhe}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_detalhe}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_detalhe, start_column=1, end_row=linha_detalhe, end_column=12)

    linha_detalhe += 1

    cabecalhos = [
        "Código",
        "Data",
        "Origem",
        "Filial",
        "Cliente",
        "Status",
        "Forma Pagamento",
        "Situação Pagamento",
        "Qtd Itens",
        "Produtos",
        "Desconto",
        "Total",
    ]

    for col, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=linha_detalhe, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    linha_atual = linha_detalhe + 1

    for item in linhas:
        valores = [
            item.get("codigo_pedido"),
            item.get("data_formatada"),
            item.get("origem_formatada"),
            item.get("unidade_nome"),
            item.get("cliente_nome"),
            item.get("status"),
            item.get("forma_pagamento"),
            item.get("situacao_pagamento"),
            item.get("qtd_itens"),
            item.get("valor_produtos"),
            item.get("valor_desconto"),
            item.get("valor_total"),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha_atual, column=col, value=valor)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")

            if col in [10, 11, 12]:
                celula.number_format = 'R$ #,##0.00'

        linha_atual += 1

    for col in range(1, 13):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 18

    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 24

    ws.freeze_panes = "A2"

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    nome_arquivo = f"fechamento_caixa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )