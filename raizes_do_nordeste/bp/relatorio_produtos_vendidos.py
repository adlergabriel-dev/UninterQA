from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.relatorio_produtos_vendidos_service import RelatorioProdutosVendidosService


bp = Blueprint(
    "relatorio_produtos_vendidos",
    __name__,
    url_prefix="/relatorios/produtos-vendidos"
)


def moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def _checkbox_ativo(valor):
    return str(valor or "").strip().lower() in ["1", "true", "on", "sim", "s"]


def _obter_filtros():
    ignorar_cancelados_raw = request.args.get("ignorar_cancelados", "1")

    return {
        "data_inicio": request.args.get("data_inicio", "").strip(),
        "data_fim": request.args.get("data_fim", "").strip(),
        "unidade_id": request.args.get("unidade_id", "").strip(),
        "produto": request.args.get("produto", "").strip(),
        "origem": request.args.get("origem", "").strip(),
        "ignorar_cancelados": _checkbox_ativo(ignorar_cancelados_raw),
    }


@bp.app_template_filter("moeda")
def filtro_moeda(valor):
    return moeda(valor)


@bp.route("/", methods=["GET"])
def index():
    filtros = _obter_filtros()

    unidades = RelatorioProdutosVendidosService.listar_unidades()

    resultado = RelatorioProdutosVendidosService.gerar_relatorio(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        produto=filtros["produto"],
        origem=filtros["origem"],
        ignorar_cancelados=filtros["ignorar_cancelados"],
    )

    return render_template(
        "relatorios/produtos_vendidos.html",
        filtros=filtros,
        unidades=unidades,
        linhas=resultado["linhas"],
        linhas_detalhe=resultado["linhas_detalhe"],
        totais=resultado["totais"],
    )


@bp.route("/exportar-excel", methods=["GET"])
def exportar_excel():
    filtros = _obter_filtros()

    resultado = RelatorioProdutosVendidosService.gerar_relatorio(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        produto=filtros["produto"],
        origem=filtros["origem"],
        ignorar_cancelados=filtros["ignorar_cancelados"],
    )

    linhas = resultado["linhas"]
    linhas_detalhe = resultado["linhas_detalhe"]
    totais = resultado["totais"]

    wb = Workbook()

    ws = wb.active
    ws.title = "Produtos Vendidos"

    cor_titulo = "1F4E78"
    cor_header = "D9EAF7"
    cor_total = "E2F0D9"
    branco = "FFFFFF"

    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "Relatório de Produtos Mais Vendidos"
    ws["A1"].font = Font(size=16, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Gerado em:"
    ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    ws["A4"] = "Período:"
    ws["B4"] = f"{filtros.get('data_inicio') or 'Início'} até {filtros.get('data_fim') or 'Hoje'}"

    ws["A5"] = "Produto:"
    ws["B5"] = filtros.get("produto") or "Todos"

    ws["D5"] = "Origem:"
    ws["E5"] = filtros.get("origem") or "Todas"

    ws["G5"] = "Ignorar Cancelados:"
    ws["H5"] = "Sim" if filtros.get("ignorar_cancelados") else "Não"

    linha_resumo = 7

    resumo = [
        ("Produtos Diferentes", totais["total_produtos_diferentes"]),
        ("Quantidade Vendida", totais["total_quantidade"]),
        ("Valor Vendido", totais["total_vendido"]),
        ("Lançamentos", totais["total_lancamentos"]),
        ("Valor Médio por Unidade", totais["ticket_medio_produto"]),
    ]

    ws[f"A{linha_resumo}"] = "Resumo Consolidado"
    ws[f"A{linha_resumo}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_resumo}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_resumo, start_column=1, end_row=linha_resumo, end_column=4)

    linha_resumo += 1

    for nome, valor in resumo:
        celula_nome = ws.cell(row=linha_resumo, column=1, value=nome)
        celula_valor = ws.cell(row=linha_resumo, column=2, value=valor)

        celula_nome.font = Font(bold=True)
        celula_nome.fill = PatternFill("solid", fgColor=cor_total)
        celula_valor.fill = PatternFill("solid", fgColor=cor_total)

        celula_nome.border = borda
        celula_valor.border = borda

        if nome in ["Valor Vendido", "Valor Médio por Unidade"]:
            celula_valor.number_format = 'R$ #,##0.00'

        linha_resumo += 1

    linha_tabela = linha_resumo + 2

    ws[f"A{linha_tabela}"] = "Produtos Agrupados"
    ws[f"A{linha_tabela}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_tabela}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_tabela, start_column=1, end_row=linha_tabela, end_column=8)

    linha_tabela += 1

    cabecalhos = [
        "Produto",
        "Origem",
        "Filial",
        "Quantidade",
        "Pedidos",
        "Valor Médio",
        "Valor Total",
        "Produto ID",
    ]

    for col, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=linha_tabela, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    linha_atual = linha_tabela + 1

    for item in linhas:
        valores = [
            item.get("produto_nome"),
            item.get("origem_formatada"),
            item.get("unidade_nome"),
            item.get("quantidade"),
            item.get("pedidos"),
            item.get("valor_unitario_medio"),
            item.get("valor_total"),
            item.get("produto_id"),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha_atual, column=col, value=valor)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")

            if col in [6, 7]:
                celula.number_format = 'R$ #,##0.00'

        linha_atual += 1

    ws_detalhe = wb.create_sheet("Detalhe dos Itens")

    ws_detalhe.merge_cells("A1:J1")
    ws_detalhe["A1"] = "Detalhamento dos Produtos Vendidos"
    ws_detalhe["A1"].font = Font(size=16, bold=True, color=branco)
    ws_detalhe["A1"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws_detalhe["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_detalhe.row_dimensions[1].height = 28

    linha_detalhe = 3

    cabecalhos_detalhe = [
        "Pedido",
        "Data",
        "Origem",
        "Cliente",
        "Filial",
        "Produto",
        "Quantidade",
        "Valor Unitário",
        "Valor Total",
        "Produto ID",
    ]

    for col, titulo in enumerate(cabecalhos_detalhe, start=1):
        celula = ws_detalhe.cell(row=linha_detalhe, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    linha_detalhe += 1

    for item in linhas_detalhe:
        valores = [
            item.get("codigo_pedido"),
            item.get("data_formatada"),
            item.get("origem_formatada"),
            item.get("cliente_nome"),
            item.get("unidade_nome"),
            item.get("produto_nome"),
            item.get("quantidade"),
            item.get("valor_unitario"),
            item.get("valor_total"),
            item.get("produto_id"),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws_detalhe.cell(row=linha_detalhe, column=col, value=valor)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")

            if col in [8, 9]:
                celula.number_format = 'R$ #,##0.00'

        linha_detalhe += 1

    for aba in [ws, ws_detalhe]:
        for col in range(1, 11):
            letra = get_column_letter(col)
            aba.column_dimensions[letra].width = 18

        aba.column_dimensions["A"].width = 28
        aba.column_dimensions["C"].width = 24
        aba.column_dimensions["D"].width = 28
        aba.column_dimensions["F"].width = 35

    ws.freeze_panes = "A2"
    ws_detalhe.freeze_panes = "A4"

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    nome_arquivo = f"relatorio_produtos_vendidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )