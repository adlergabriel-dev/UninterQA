from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.relatorio_pedido_balcao_service import RelatorioPedidoBalcaoService


bp = Blueprint(
    "relatorio_pedido_balcao",
    __name__,
    url_prefix="/relatorios/pedido-balcao"
)


def moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def _obter_filtros():
    return {
        "data_inicio": request.args.get("data_inicio", "").strip(),
        "data_fim": request.args.get("data_fim", "").strip(),
        "unidade_id": request.args.get("unidade_id", "").strip(),
        "status": request.args.get("status", "").strip(),
        "forma_pagamento": request.args.get("forma_pagamento", "").strip(),
        "situacao_pagamento": request.args.get("situacao_pagamento", "").strip(),
    }


@bp.app_template_filter("moeda")
def filtro_moeda(valor):
    return moeda(valor)


@bp.route("/", methods=["GET"])
def index():
    filtros = _obter_filtros()

    unidades = RelatorioPedidoBalcaoService.listar_unidades()

    resultado = RelatorioPedidoBalcaoService.gerar_relatorio(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        status=filtros["status"],
        forma_pagamento=filtros["forma_pagamento"],
        situacao_pagamento=filtros["situacao_pagamento"],
    )

    return render_template(
        "relatorios/pedido_balcao.html",
        filtros=filtros,
        unidades=unidades,
        linhas=resultado["linhas"],
        totais=resultado["totais"],
    )


@bp.route("/exportar-excel", methods=["GET"])
def exportar_excel():
    filtros = _obter_filtros()

    resultado = RelatorioPedidoBalcaoService.gerar_relatorio(
        data_inicio=filtros["data_inicio"],
        data_fim=filtros["data_fim"],
        unidade_id=filtros["unidade_id"],
        status=filtros["status"],
        forma_pagamento=filtros["forma_pagamento"],
        situacao_pagamento=filtros["situacao_pagamento"],
    )

    linhas = resultado["linhas"]
    totais = resultado["totais"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido Balcão"

    cor_titulo = "1F4E78"
    cor_header = "D9EAF7"
    cor_total = "E2F0D9"
    branco = "FFFFFF"

    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:N1")
    ws["A1"] = "Relatório de Pedido Balcão"
    ws["A1"].font = Font(size=16, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Gerado em:"
    ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    ws["A4"] = "Período:"
    ws["B4"] = f"{filtros.get('data_inicio') or 'Início'} até {filtros.get('data_fim') or 'Hoje'}"

    ws["A5"] = "Status:"
    ws["B5"] = filtros.get("status") or "Todos"

    ws["D5"] = "Forma Pagamento:"
    ws["E5"] = filtros.get("forma_pagamento") or "Todas"

    ws["G5"] = "Situação Pagamento:"
    ws["H5"] = filtros.get("situacao_pagamento") or "Todas"

    linha_resumo = 7

    resumo = [
        ("Total de Pedidos", totais["total_pedidos"]),
        ("Total de Itens", totais["total_itens"]),
        ("Valor Produtos", totais["total_produtos"]),
        ("Descontos", totais["total_descontos"]),
        ("Entrega", totais["total_entregas"]),
        ("Total Geral", totais["total_geral"]),
        ("Total Pago", totais["total_pago"]),
        ("Total Pendente", totais["total_pendente"]),
        ("Total Cancelado", totais["total_cancelado"]),
    ]

    ws[f"A{linha_resumo}"] = "Resumo Consolidado"
    ws[f"A{linha_resumo}"].font = Font(bold=True, color=branco)
    ws[f"A{linha_resumo}"].fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=linha_resumo, start_column=1, end_row=linha_resumo, end_column=4)

    linha_resumo += 1

    for nome, valor in resumo:
        celula_nome = ws.cell(row=linha_resumo, column=1, value=nome)
        celula_valor = ws.cell(row=linha_resumo, column=2, value=valor)

        celula_nome.font = Font(bold=True)
        celula_nome.fill = PatternFill("solid", fgColor=cor_total)
        celula_valor.fill = PatternFill("solid", fgColor=cor_total)

        celula_nome.border = borda
        celula_valor.border = borda

        if nome not in ["Total de Pedidos", "Total de Itens"]:
            celula_valor.number_format = 'R$ #,##0.00'

        linha_resumo += 1

    linha_tabela = linha_resumo + 2

    cabecalhos = [
        "Código",
        "Data",
        "Cliente",
        "Telefone",
        "Filial",
        "Status",
        "Forma Pagamento",
        "Situação Pagamento",
        "Qtd Itens",
        "Valor Produtos",
        "Desconto",
        "Entrega",
        "Total",
        "Observação",
    ]

    for col, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=linha_tabela, column=col, value=titulo)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor=cor_header)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    linha_atual = linha_tabela + 1

    for item in linhas:
        valores = [
            item.get("codigo_pedido"),
            item.get("data_formatada"),
            item.get("cliente_nome"),
            item.get("cliente_telefone"),
            item.get("unidade_nome"),
            item.get("status"),
            item.get("forma_pagamento"),
            item.get("situacao_pagamento"),
            item.get("qtd_itens"),
            item.get("valor_produtos"),
            item.get("valor_desconto"),
            item.get("valor_entrega"),
            item.get("valor_total"),
            item.get("observacao"),
        ]

        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha_atual, column=col, value=valor)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")

            if col in [10, 11, 12, 13]:
                celula.number_format = 'R$ #,##0.00'

        linha_atual += 1

    for col in range(1, len(cabecalhos) + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 18

    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["N"].width = 40

    ws.freeze_panes = ws.cell(row=linha_tabela + 1, column=1)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    nome_arquivo = f"relatorio_pedido_balcao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )